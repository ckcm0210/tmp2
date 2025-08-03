#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
統一INDIRECT解析器
支援兩種模式：
1. Excel模式 - 打開Excel，在Excel中計算INDIRECT
2. Pure模式 - 不打開Excel，純openpyxl解析
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sys
import os
import time
import re
import stat
from urllib.parse import unquote
sys.path.append('.')

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import win32com.client as win32
    HAS_WIN32COM = True
except ImportError:
    HAS_WIN32COM = False

class UnifiedIndirectResolver:
    """統一INDIRECT解析器"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Unified INDIRECT Resolver - Excel & Pure Modes")
        self.root.geometry("1200x800")
        
        # 狀態變數
        self.current_file = None
        self.workbook = None
        self.worksheet = None
        self.xl = None
        self.excel_workbook = None
        self.external_links_map = {}
        self.external_workbooks = {}
        
        self.setup_ui()
        
        if not HAS_OPENPYXL:
            messagebox.showerror("Error", "Need openpyxl: pip install openpyxl")
    
    def setup_ui(self):
        """設置UI"""
        
        # 標題
        title_label = tk.Label(self.root, text="Unified INDIRECT Resolver", 
                              font=("Arial", 16, "bold"))
        title_label.pack(pady=10)
        
        # 模式選擇
        mode_frame = ttk.LabelFrame(self.root, text="Resolution Mode", padding=10)
        mode_frame.pack(pady=10, padx=10, fill="x")
        
        self.mode_var = tk.StringVar(value="excel")
        
        excel_radio = ttk.Radiobutton(mode_frame, text="Excel Mode (Fast & Accurate)", 
                                     variable=self.mode_var, value="excel")
        excel_radio.pack(side=tk.LEFT, padx=10)
        
        pure_radio = ttk.Radiobutton(mode_frame, text="Pure Mode (No Excel Required)", 
                                    variable=self.mode_var, value="pure")
        pure_radio.pack(side=tk.LEFT, padx=10)
        
        # 模式說明
        mode_desc = tk.Label(mode_frame, 
                           text="Excel Mode: Uses Excel COM for calculation (requires Excel)\n" +
                                "Pure Mode: Uses openpyxl only (works without Excel)",
                           font=("Arial", 9), fg="gray")
        mode_desc.pack(pady=5)
        
        # 輸入區域
        input_frame = ttk.LabelFrame(self.root, text="Input", padding=10)
        input_frame.pack(pady=10, padx=10, fill="x")
        
        # 文件選擇
        file_frame = ttk.Frame(input_frame)
        file_frame.pack(fill="x", pady=5)
        
        ttk.Label(file_frame, text="Excel File:").pack(side=tk.LEFT)
        self.file_path_var = tk.StringVar()
        file_entry = ttk.Entry(file_frame, textvariable=self.file_path_var, width=50)
        file_entry.pack(side=tk.LEFT, padx=5, fill="x", expand=True)
        ttk.Button(file_frame, text="Browse", command=self.browse_file).pack(side=tk.LEFT, padx=5)
        
        # 工作表選擇
        sheet_frame = ttk.Frame(input_frame)
        sheet_frame.pack(fill="x", pady=5)
        
        ttk.Label(sheet_frame, text="Worksheet:").pack(side=tk.LEFT)
        self.sheet_var = tk.StringVar()
        self.sheet_combo = ttk.Combobox(sheet_frame, textvariable=self.sheet_var, width=30)
        self.sheet_combo.pack(side=tk.LEFT, padx=5)
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_selected)
        
        # 儲存格地址
        cell_frame = ttk.Frame(input_frame)
        cell_frame.pack(fill="x", pady=5)
        
        ttk.Label(cell_frame, text="Cell Address:").pack(side=tk.LEFT)
        self.cell_var = tk.StringVar()
        cell_entry = ttk.Entry(cell_frame, textvariable=self.cell_var, width=20)
        cell_entry.pack(side=tk.LEFT, padx=5)
        
        # 主要按鈕
        button_frame = ttk.Frame(input_frame)
        button_frame.pack(fill="x", pady=10)
        
        resolve_btn = ttk.Button(button_frame, text="Resolve INDIRECT", 
                                command=self.resolve_indirect_unified, width=20)
        resolve_btn.pack(side=tk.LEFT, padx=5)
        
        clear_btn = ttk.Button(button_frame, text="Clear Results", 
                              command=self.clear_results, width=15)
        clear_btn.pack(side=tk.LEFT, padx=5)
        
        # 結果顯示
        result_frame = ttk.LabelFrame(self.root, text="Results", padding=10)
        result_frame.pack(pady=10, padx=10, fill="both", expand=True)
        
        text_container = ttk.Frame(result_frame)
        text_container.pack(fill="both", expand=True)
        
        self.result_text = tk.Text(text_container, wrap=tk.WORD, font=("Consolas", 10))
        scrollbar = ttk.Scrollbar(text_container, orient="vertical", command=self.result_text.yview)
        self.result_text.configure(yscrollcommand=scrollbar.set)
        
        self.result_text.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
    def browse_file(self):
        """瀏覽文件"""
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.file_path_var.set(file_path)
            self.load_file_info()
    
    def load_file_info(self):
        """載入文件信息"""
        file_path = self.file_path_var.get()
        if not file_path or not os.path.exists(file_path):
            return
        
        try:
            # 載入工作簿
            self.workbook = openpyxl.load_workbook(file_path, data_only=False)
            self.current_file = file_path
            
            self.add_result(f"File loaded: {os.path.basename(file_path)}")
            
            # 獲取外部連結映射
            self.get_external_links_from_openpyxl()
            
            # 更新工作表列表
            self.sheet_combo['values'] = self.workbook.sheetnames
            if self.workbook.sheetnames:
                self.sheet_combo.current(0)
                self.on_sheet_selected(None)
            
        except Exception as e:
            self.add_result(f"Error loading file: {str(e)}")
            messagebox.showerror("Error", f"Cannot load file: {str(e)}")
    
    def on_sheet_selected(self, event):
        """工作表選擇事件"""
        if not self.workbook:
            return
        
        sheet_name = self.sheet_var.get()
        if sheet_name:
            try:
                self.worksheet = self.workbook[sheet_name]
                self.add_result(f"Selected worksheet: {sheet_name}")
            except Exception as e:
                self.add_result(f"Error selecting worksheet: {str(e)}")
    
    def resolve_indirect_unified(self):
        """統一的INDIRECT解析入口"""
        if not self.validate_inputs():
            return
        
        mode = self.mode_var.get()
        
        self.add_result("=" * 60)
        self.add_result(f"RESOLVING INDIRECT - {mode.upper()} MODE")
        self.add_result("=" * 60)
        
        if mode == "excel":
            self.resolve_with_excel_mode()
        else:
            self.resolve_with_pure_mode()
    
    def resolve_with_excel_mode(self):
        """Excel模式解析"""
        try:
            self.add_result("Using Excel Mode - Fast & Accurate")
            
            # 連接Excel
            if not self.connect_to_excel():
                return
            
            # 獲取原始公式
            cell_address = self.cell_var.get().upper()
            sheet_name = self.sheet_var.get()
            
            # 從Excel獲取公式
            excel_ws = self.excel_workbook.Worksheets(sheet_name)
            excel_cell = excel_ws.Range(cell_address)
            formula = excel_cell.Formula
            
            self.add_result(f"File: {os.path.basename(self.current_file)}")
            self.add_result(f"Worksheet: {sheet_name}")
            self.add_result(f"Cell: {cell_address}")
            self.add_result(f"Formula: {formula}")
            
            if not formula or "INDIRECT" not in formula.upper():
                self.add_result("No INDIRECT function found")
                return
            
            # 使用安全的Excel計算方法
            result = self.safe_excel_calculation(excel_cell, formula)
            
            if result is not None:
                self.add_result(f"\nEXCEL MODE RESULT:")
                self.add_result(f"INDIRECT resolves to: {result}")
            else:
                self.add_result("Excel mode calculation failed")
                
        except Exception as e:
            self.add_result(f"Error in Excel mode: {str(e)}")
    
    def resolve_with_pure_mode(self):
        """Pure模式解析"""
        try:
            self.add_result("Using Pure Mode - No Excel Required")
            
            # 獲取原始公式
            cell_address = self.cell_var.get().upper()
            sheet_name = self.sheet_var.get()
            
            cell = self.worksheet[cell_address]
            original_formula = cell.value
            
            self.add_result(f"File: {os.path.basename(self.current_file)}")
            self.add_result(f"Worksheet: {sheet_name}")
            self.add_result(f"Cell: {cell_address}")
            self.add_result(f"Formula: {original_formula}")
            
            # 處理ArrayFormula對象
            formula_text = self.extract_formula_text(original_formula)
            if not formula_text or "INDIRECT" not in formula_text.upper():
                self.add_result("No INDIRECT function found")
                return
            
            # 使用Pure模式解析
            result = self.pure_mode_calculation(formula_text)
            
            if result:
                self.add_result(f"\nPURE MODE RESULT:")
                self.add_result(f"INDIRECT resolves to: {result}")
            else:
                self.add_result("Pure mode calculation failed")
                
        except Exception as e:
            self.add_result(f"Error in Pure mode: {str(e)}")
    
    def connect_to_excel(self):
        """連接Excel"""
        if not HAS_WIN32COM:
            self.add_result("Excel COM not available")
            return False
        
        try:
            self.add_result("Connecting to Excel...")
            self.xl = win32.GetActiveObject("Excel.Application")
            self.excel_workbook = self.xl.Workbooks.Open(self.current_file)
            self.add_result("Successfully connected to Excel")
            return True
        except Exception as e:
            self.add_result(f"Failed to connect to Excel: {str(e)}")
            return False
    
    def safe_excel_calculation(self, excel_cell, formula):
        """安全的Excel計算方法"""
        try:
            self.add_result("Starting safe Excel calculation...")
            
            # 保存原始狀態
            original_formula = excel_cell.Formula
            original_calculation = self.xl.Calculation
            original_events = self.xl.EnableEvents
            original_screen_updating = self.xl.ScreenUpdating
            original_interactive = self.xl.Interactive
            
            # 設置保護模式
            self.xl.Calculation = -4135  # xlCalculationManual
            self.xl.EnableEvents = False
            self.xl.ScreenUpdating = False
            self.xl.Interactive = False
            
            self.add_result("Protection mode activated")
            
            try:
                # 提取INDIRECT內容
                indirect_content = self.extract_indirect_content_excel(formula)
                if not indirect_content:
                    return None
                
                self.add_result(f"INDIRECT content: {indirect_content}")
                
                # 在Excel中計算
                excel_cell.Formula = f"={indirect_content}"
                excel_cell.Calculate()
                result = excel_cell.Value
                
                self.add_result(f"Excel calculation result: {result}")
                return result
                
            finally:
                # 恢復所有狀態
                excel_cell.Formula = original_formula
                self.xl.Calculation = original_calculation
                self.xl.EnableEvents = original_events
                self.xl.ScreenUpdating = original_screen_updating
                self.xl.Interactive = original_interactive
                
                self.add_result("All states restored")
                
        except Exception as e:
            self.add_result(f"Error in safe Excel calculation: {str(e)}")
            return None
    
    def pure_mode_calculation(self, formula):
        """Pure模式計算 - 使用完整的跨文件解析邏輯"""
        try:
            self.add_result("Starting pure mode calculation...")
            
            # 提取INDIRECT內容
            indirect_content = self.extract_indirect_content_pure(formula)
            if not indirect_content:
                return None
            
            self.add_result(f"INDIRECT content: {indirect_content}")
            
            # 修復外部引用
            fixed_content = self.fix_external_references(indirect_content)
            self.add_result(f"After external reference fix: {fixed_content}")
            
            # 解析字串連接
            if '&' in fixed_content:
                self.add_result("Contains string concatenation - resolving components:")
                components = self.resolve_concatenation_components_full(fixed_content)
                result = self.build_final_reference_full(components)
                return result
            else:
                self.add_result("Simple reference - no concatenation")
                return fixed_content
                
        except Exception as e:
            self.add_result(f"Error in pure mode calculation: {str(e)}")
            return None
    
    def extract_indirect_content_excel(self, formula):
        """Excel模式提取INDIRECT內容 - 使用括號配對，處理引號"""
        try:
            self.add_result(f"Extracting INDIRECT from: {formula}")
            
            # 找到INDIRECT的開始位置
            indirect_start = formula.upper().find('INDIRECT(')
            if indirect_start == -1:
                return None
            
            # 從INDIRECT(後開始計算括號
            start_pos = indirect_start + len('INDIRECT(')
            bracket_count = 1
            current_pos = start_pos
            in_quotes = False
            quote_char = None
            
            # 逐字符掃描，計算括號配對，處理引號
            while current_pos < len(formula) and bracket_count > 0:
                char = formula[current_pos]
                
                # 處理引號（單引號和雙引號）
                if char in ['"', "'"] and not in_quotes:
                    in_quotes = True
                    quote_char = char
                elif char == quote_char and in_quotes:
                    in_quotes = False
                    quote_char = None
                
                # 只在不在引號內時計算括號
                if not in_quotes:
                    if char == '(':
                        bracket_count += 1
                    elif char == ')':
                        bracket_count -= 1
                
                current_pos += 1
            
            if bracket_count == 0:
                content = formula[start_pos:current_pos-1]
                self.add_result(f"INDIRECT content extracted: {content}")
                return content
            else:
                self.add_result(f"Unmatched brackets - bracket_count: {bracket_count}")
                return None
            
        except Exception as e:
            self.add_result(f"Error extracting INDIRECT: {str(e)}")
            return None
    
    def extract_indirect_content_pure(self, formula):
        """Pure模式提取INDIRECT內容 - 處理嵌套括號和引號"""
        try:
            indirect_start = formula.upper().find('INDIRECT(')
            if indirect_start == -1:
                return None
            
            start_pos = indirect_start + len('INDIRECT(')
            bracket_count = 1
            current_pos = start_pos
            in_quotes = False
            quote_char = None
            
            while current_pos < len(formula) and bracket_count > 0:
                char = formula[current_pos]
                
                # 處理引號（單引號和雙引號）
                if char in ['"', "'"] and not in_quotes:
                    in_quotes = True
                    quote_char = char
                elif char == quote_char and in_quotes:
                    in_quotes = False
                    quote_char = None
                
                # 只在不在引號內時計算括號
                if not in_quotes:
                    if char == '(':
                        bracket_count += 1
                    elif char == ')':
                        bracket_count -= 1
                
                current_pos += 1
            
            if bracket_count == 0:
                return formula[start_pos:current_pos-1]
            
            return None
        except:
            return None
    
    def extract_formula_text(self, formula_obj):
        """從不同類型的公式對象中提取文字"""
        try:
            if formula_obj is None:
                return None
            
            if isinstance(formula_obj, str):
                return formula_obj
            
            if hasattr(formula_obj, 'text'):
                return formula_obj.text
            
            return str(formula_obj)
        except:
            return None
    
    def get_external_links_from_openpyxl(self):
        """從openpyxl獲取外部連結映射"""
        try:
            self.external_links_map = {}
            
            if hasattr(self.workbook, '_external_links'):
                external_links = self.workbook._external_links
                if external_links:
                    for i, link in enumerate(external_links, 1):
                        if hasattr(link, 'file_link') and link.file_link:
                            file_path = link.file_link.Target
                            if file_path:
                                decoded_path = unquote(file_path)
                                if decoded_path.startswith('file:///'):
                                    decoded_path = decoded_path[8:]
                                elif decoded_path.startswith('file://'):
                                    decoded_path = decoded_path[7:]
                                
                                self.external_links_map[str(i)] = decoded_path
                                filename = os.path.basename(decoded_path)
                                self.add_result(f"External link [{i}] = {filename}")
            
            if not self.external_links_map:
                self.infer_external_links_from_formulas()
                
        except Exception as e:
            self.infer_external_links_from_formulas()
    
    def infer_external_links_from_formulas(self):
        """從公式中推斷外部連結"""
        try:
            base_dir = os.path.dirname(self.current_file)
            common_files = [
                "Link1.xlsx", "Link2.xlsx", "Link3.xlsx",
                "File1.xlsx", "File2.xlsx", "File3.xlsx",
                "Data.xlsx", "GDP.xlsx", "Test.xlsx"
            ]
            
            index = 1
            for filename in common_files:
                full_path = os.path.join(base_dir, filename)
                if os.path.exists(full_path):
                    self.external_links_map[str(index)] = full_path
                    self.add_result(f"Inferred link [{index}] = {filename}")
                    index += 1
                    
        except Exception as e:
            self.add_result(f"Error inferring external links: {str(e)}")
    
    def fix_external_references(self, content):
        """修復外部引用"""
        try:
            def replace_ref(match):
                ref_num = match.group(1)
                if ref_num in self.external_links_map:
                    full_path = self.external_links_map[ref_num]
                    decoded_path = unquote(full_path) if isinstance(full_path, str) else full_path
                    if decoded_path.startswith('file:///'):
                        decoded_path = decoded_path[8:]
                    
                    filename = os.path.basename(decoded_path)
                    directory = os.path.dirname(decoded_path)
                    return f"'{directory}\\[{filename}]'"
                return f"[Unknown_{ref_num}]"
            
            pattern = r'\[(\d+)\]'
            return re.sub(pattern, replace_ref, content)
        except:
            return content
    
    def resolve_concatenation_components_full(self, content):
        """解析字串連接組件 - 智能分析版本"""
        try:
            self.add_result("Starting intelligent component analysis...")
            
            # 第一步：智能分割（處理引號入面嘅&）
            parts = self.smart_split_by_ampersand(content)
            self.add_result(f"Split into {len(parts)} parts:")
            for i, part in enumerate(parts):
                self.add_result(f"  Part {i+1}: {part}")
            
            components = []
            
            # 第二步：分析每個組件
            for part in parts:
                part = part.strip()
                self.add_result(f"\nAnalyzing component: {part}")
                
                # 識別組件類型
                comp_type, comp_data = self.identify_component_type(part)
                self.add_result(f"  Type: {comp_type}")
                
                if comp_type == 'string':
                    components.append(('string', comp_data))
                    self.add_result(f"  String constant: '{comp_data}'")
                
                elif comp_type == 'cell':
                    cell_value = self.get_cell_value_with_formula_calc(comp_data)
                    components.append(('cell', comp_data, cell_value))
                    self.add_result(f"  Cell {comp_data} = {cell_value}")
                
                elif comp_type == 'function':
                    # 提取完整函數（處理嵌套括弧）
                    complete_func = self.extract_complete_function_from_part(part, content)
                    # 傳入正確嘅儲存格地址作為context
                    func_result = self.resolve_function_smart(complete_func, comp_data)
                    components.append(('function', complete_func, func_result))
                    self.add_result(f"  Function result = {func_result}")
                
                else:
                    # 其他表達式
                    components.append(('expression', part, None))
                    self.add_result(f"  Complex expression: {part}")
            
            return components
        except Exception as e:
            self.add_result(f"Error in intelligent analysis: {str(e)}")
            return []
    
    def build_final_reference_full(self, components):
        """構建最終引用 - 完整版"""
        try:
            self.add_result("\nBuilding final reference from components:")
            
            result_parts = []
            for component in components:
                comp_type = component[0]
                comp_value = component[2] if len(component) > 2 else None
                
                if comp_type == 'string':
                    result_parts.append(component[1])
                elif comp_type in ['cell', 'vlookup', 'external'] and comp_value is not None:
                    result_parts.append(str(comp_value))
                else:
                    result_parts.append(f"({component[1]})")
            
            final_reference = ''.join(result_parts)
            self.add_result(f"Final reference: {final_reference}")
            
            return final_reference
        except Exception as e:
            self.add_result(f"Error building final reference: {str(e)}")
            return None
    
    def get_cell_value_with_formula_calc(self, cell_ref):
        """獲取儲存格值，包含公式計算 - 增強版本"""
        try:
            cell = self.worksheet[cell_ref]
            raw_value = cell.value
            
            # 如果唔係公式，直接返回
            if not (isinstance(raw_value, str) and raw_value.startswith('=')):
                return raw_value
            
            # 係公式，嘗試計算
            self.add_result(f"    Found formula in {cell_ref}: {raw_value}")
            
            # 移除開頭嘅=號
            formula = raw_value[1:]
            
            # 嘗試計算不同類型嘅公式
            if 'SUM(' in formula.upper():
                result = self.calculate_sum_formula(raw_value)
                self.add_result(f"    SUM formula result: {result}")
                return result
            
            elif '&' in formula:
                # 字串連接公式
                result = self.calculate_string_concat_formula(formula, cell_ref)
                self.add_result(f"    String concat result: {result}")
                return result
            
            elif 'ROW()' in formula.upper():
                # ROW函數
                result = self.resolve_position_aware_function(formula, 'ROW', cell_ref)
                self.add_result(f"    ROW function result: {result}")
                return result
            
            elif 'COLUMN()' in formula.upper():
                # COLUMN函數
                result = self.resolve_position_aware_function(formula, 'COLUMN', cell_ref)
                self.add_result(f"    COLUMN function result: {result}")
                return result
            
            else:
                # 其他公式暫時返回原始值
                self.add_result(f"    Unknown formula type, returning original: {raw_value}")
                return raw_value
            
        except Exception as e:
            self.add_result(f"    Error calculating formula: {str(e)}")
            return raw_value if 'raw_value' in locals() else None
    
    def calculate_sum_formula(self, formula):
        """計算SUM公式"""
        try:
            pattern = r'SUM\s*\(\s*([A-Z]+\d+:[A-Z]+\d+)\s*\)'
            match = re.search(pattern, formula, re.IGNORECASE)
            
            if match:
                range_ref = match.group(1)
                start_cell, end_cell = range_ref.split(':')
                
                total = 0
                start_col = re.match(r'([A-Z]+)', start_cell).group(1)
                start_row = int(re.match(r'[A-Z]+(\d+)', start_cell).group(1))
                end_row = int(re.match(r'[A-Z]+(\d+)', end_cell).group(1))
                
                for row in range(start_row, end_row + 1):
                    cell_addr = f"{start_col}{row}"
                    try:
                        cell_value = self.worksheet[cell_addr].value
                        if isinstance(cell_value, (int, float)):
                            total += cell_value
                    except:
                        continue
                
                return total
            
            return None
        except:
            return None
    
    def calculate_string_concat_formula(self, formula, current_cell):
        """計算字串連接公式"""
        try:
            self.add_result(f"      Calculating string concat: {formula}")
            
            # 按 & 分割
            parts = self.smart_split_by_ampersand(formula)
            result_parts = []
            
            for part in parts:
                part = part.strip()
                self.add_result(f"      Processing part: {part}")
                
                if part.startswith('"') and part.endswith('"'):
                    # 字串常數
                    value = part[1:-1]
                    result_parts.append(value)
                    self.add_result(f"        String: '{value}'")
                
                elif re.match(r'^\$?[A-Z]+\$?\d+$', part):
                    # 儲存格引用
                    cell_value = self.worksheet[part].value
                    result_parts.append(str(cell_value) if cell_value is not None else "")
                    self.add_result(f"        Cell {part}: {cell_value}")
                
                elif 'ROW()' in part.upper():
                    # ROW函數
                    row_result = self.resolve_position_aware_function(part, 'ROW', current_cell)
                    result_parts.append(str(row_result))
                    self.add_result(f"        ROW function: {row_result}")
                
                elif 'COLUMN()' in part.upper():
                    # COLUMN函數
                    col_result = self.resolve_position_aware_function(part, 'COLUMN', current_cell)
                    result_parts.append(str(col_result))
                    self.add_result(f"        COLUMN function: {col_result}")
                
                else:
                    # 其他，保持原樣
                    result_parts.append(part)
                    self.add_result(f"        Other: {part}")
            
            final_result = ''.join(result_parts)
            self.add_result(f"      Final concat result: {final_result}")
            return final_result
            
        except Exception as e:
            self.add_result(f"      Error in string concat: {str(e)}")
            return formula
    
    def resolve_vlookup_full(self, vlookup_expr):
        """完整的VLOOKUP解析 - 修復版本"""
        try:
            self.add_result(f"    Resolving VLOOKUP: {vlookup_expr}")
            
            # 處理不完整的VLOOKUP表達式（缺少結束括號）
            if not vlookup_expr.strip().endswith(')'):
                if vlookup_expr.count(',') >= 2:
                    vlookup_expr = vlookup_expr.strip() + ',FALSE)'
                else:
                    vlookup_expr = vlookup_expr.strip() + ')'
                self.add_result(f"    Fixed expression: {vlookup_expr}")
            
            # 使用正則表達式解析VLOOKUP參數
            pattern = r'VLOOKUP\s*\(\s*([^,]+)\s*,\s*([^,]+)\s*,\s*([^,]+)\s*(?:,\s*([^)]+))?\s*\)'
            match = re.search(pattern, vlookup_expr, re.IGNORECASE)
            
            if not match:
                self.add_result(f"    Could not parse VLOOKUP: {vlookup_expr}")
                return None
            
            lookup_value = match.group(1).strip()
            table_range = match.group(2).strip()
            col_index_str = match.group(3).strip()
            match_type = match.group(4).strip() if match.group(4) else 'FALSE'
            
            self.add_result(f"    Lookup value: {lookup_value}")
            self.add_result(f"    Table range: {table_range}")
            self.add_result(f"    Column index: {col_index_str}")
            
            # 轉換列索引為整數
            try:
                col_index = int(col_index_str)
            except ValueError:
                self.add_result(f"    Invalid column index: {col_index_str}")
                return None
            
            # 處理查找值
            if lookup_value.startswith('"') and lookup_value.endswith('"'):
                lookup_value = lookup_value[1:-1]
            elif re.match(r'^[A-Z]+\d+$', lookup_value):
                # 是儲存格引用，獲取其值
                cell_value = self.get_cell_value_with_formula_calc(lookup_value)
                self.add_result(f"    Cell {lookup_value} value: {cell_value}")
                lookup_value = cell_value
            
            # 執行VLOOKUP
            result = self.perform_vlookup_simple(lookup_value, table_range, col_index)
            self.add_result(f"    VLOOKUP result: {result}")
            return result
            
        except Exception as e:
            self.add_result(f"    Error in VLOOKUP resolution: {str(e)}")
            return None
    
    def perform_vlookup_simple(self, lookup_value, table_range, col_index):
        """簡化的VLOOKUP執行 - 增強版本"""
        try:
            self.add_result(f"      Performing VLOOKUP search...")
            self.add_result(f"      Looking for: {lookup_value}")
            self.add_result(f"      In range: {table_range}")
            self.add_result(f"      Return column: {col_index}")
            
            if ':' not in table_range:
                self.add_result(f"      Invalid table range: {table_range}")
                return None
                
            start_cell, end_cell = table_range.split(':')
            
            # 解析範圍
            start_col_match = re.match(r'\$?([A-Z]+)', start_cell)
            start_row_match = re.match(r'\$?[A-Z]+\$?(\d+)', start_cell)
            end_row_match = re.match(r'\$?[A-Z]+\$?(\d+)', end_cell)
            
            if not all([start_col_match, start_row_match, end_row_match]):
                self.add_result(f"      Could not parse range: {table_range}")
                return None
            
            start_col = start_col_match.group(1)
            start_row = int(start_row_match.group(1))
            end_row = int(end_row_match.group(1))
            
            self.add_result(f"      Searching in {start_col}{start_row} to {start_col}{end_row}")
            
            # 搜尋匹配值
            for row in range(start_row, end_row + 1):
                first_col_cell = f"{start_col}{row}"
                try:
                    cell_value = self.worksheet[first_col_cell].value
                    self.add_result(f"      Checking {first_col_cell}: {cell_value}")
                    
                    if self.values_match_simple(cell_value, lookup_value):
                        # 找到匹配，返回指定列的值
                        result_col = chr(ord(start_col) + col_index - 1)
                        result_cell = f"{result_col}{row}"
                        result_value = self.worksheet[result_cell].value
                        self.add_result(f"      Match found! Returning {result_cell}: {result_value}")
                        return result_value
                except Exception as e:
                    self.add_result(f"      Error checking {first_col_cell}: {str(e)}")
                    continue
            
            self.add_result(f"      No match found for: {lookup_value}")
            return None
        except Exception as e:
            self.add_result(f"      Error in VLOOKUP execution: {str(e)}")
            return None
    
    def values_match_simple(self, cell_value, lookup_value):
        """簡化的值匹配"""
        try:
            if cell_value is None:
                return False
            
            # 數字比較
            if isinstance(cell_value, (int, float)) and isinstance(lookup_value, (int, float)):
                return cell_value == lookup_value
            
            # 字串比較
            return str(cell_value).strip().upper() == str(lookup_value).strip().upper()
        except:
            return False
    
    def smart_split_by_ampersand(self, content):
        """按 & 分割，但唔會分割引號入面嘅 &"""
        try:
            parts = []
            current_part = ""
            in_quotes = False
            quote_char = None
            
            i = 0
            while i < len(content):
                char = content[i]
                
                # 處理引號
                if char in ['"', "'"] and not in_quotes:
                    in_quotes = True
                    quote_char = char
                    current_part += char
                elif char == quote_char and in_quotes:
                    in_quotes = False
                    quote_char = None
                    current_part += char
                elif char == '&' and not in_quotes:
                    # 分割點
                    if current_part.strip():
                        parts.append(current_part.strip())
                    current_part = ""
                else:
                    current_part += char
                
                i += 1
            
            # 加最後一部分
            if current_part.strip():
                parts.append(current_part.strip())
            
            return parts
        except Exception as e:
            self.add_result(f"Error in smart split: {str(e)}")
            return [content]
    
    def identify_component_type(self, component):
        """識別組件類型"""
        try:
            comp = component.strip()
            
            # 1. 字串常數 (用引號包住)
            if (comp.startswith('"') and comp.endswith('"')) or \
               (comp.startswith("'") and comp.endswith("'")):
                return ('string', comp[1:-1])
            
            # 2. 簡單儲存格引用 (例如: B8, $A$1)
            if re.match(r'^\$?[A-Z]+\$?\d+$', comp):
                return ('cell', comp)
            
            # 3. 函數 (有英文字母 + 開括弧)
            if re.match(r'^[A-Z]+\s*\(', comp):
                return ('function', comp)
            
            # 4. 其他表達式
            return ('expression', comp)
        except:
            return ('expression', component)
    
    def extract_complete_function_from_part(self, part, full_content):
        """從部分提取完整函數"""
        try:
            # 如果part已經係完整函數，直接返回
            if part.count('(') == part.count(')') and part.count('(') > 0:
                return part
            
            # 否則喺full_content入面搵完整函數
            func_start = full_content.find(part)
            if func_start == -1:
                return part
            
            return self.extract_complete_function(part, full_content, func_start)
        except:
            return part
    
    def extract_complete_function(self, func_start, full_content, start_pos):
        """從函數開始位置提取完整函數（包括嵌套括弧）"""
        try:
            bracket_count = 0
            current_pos = start_pos
            in_quotes = False
            quote_char = None
            
            # 找到第一個開括弧
            while current_pos < len(full_content):
                if full_content[current_pos] == '(':
                    bracket_count = 1
                    current_pos += 1
                    break
                current_pos += 1
            
            if bracket_count == 0:
                return func_start
            
            # 繼續掃描直到括弧配對完成
            while current_pos < len(full_content) and bracket_count > 0:
                char = full_content[current_pos]
                
                # 處理引號
                if char in ['"', "'"] and not in_quotes:
                    in_quotes = True
                    quote_char = char
                elif char == quote_char and in_quotes:
                    in_quotes = False
                    quote_char = None
                
                # 只在引號外計算括弧
                if not in_quotes:
                    if char == '(':
                        bracket_count += 1
                    elif char == ')':
                        bracket_count -= 1
                
                current_pos += 1
            
            return full_content[start_pos:current_pos]
        except:
            return func_start
    
    def resolve_function_smart(self, function_text, context_cell=None):
        """智能解析函數"""
        try:
            func_upper = function_text.upper()
            
            # VLOOKUP函數
            if func_upper.startswith('VLOOKUP'):
                return self.resolve_vlookup_full(function_text)
            
            # ROW函數
            elif func_upper.startswith('ROW'):
                return self.resolve_position_aware_function(function_text, 'ROW', context_cell)
            
            # COLUMN函數
            elif func_upper.startswith('COLUMN'):
                return self.resolve_position_aware_function(function_text, 'COLUMN', context_cell)
            
            # 其他函數暫時返回原始文字
            else:
                self.add_result(f"    Unknown function type: {function_text}")
                return function_text
                
        except Exception as e:
            self.add_result(f"    Error resolving function: {str(e)}")
            return function_text
    
    def resolve_position_aware_function(self, func_expr, func_type, context_cell=None):
        """解析位置相關函數（ROW, COLUMN等）"""
        try:
            # 使用context_cell（公式實際所在嘅儲存格）而唔係當前選中嘅儲存格
            target_cell = context_cell if context_cell else self.cell_var.get()
            
            if target_cell and func_type.upper() == 'ROW':
                row_num = int(re.search(r'\d+', target_cell).group())
                self.add_result(f"        ROW() context: {target_cell}, row: {row_num}")
                
                # 處理 ROW()+數字 嘅情況
                if '+' in func_expr:
                    match = re.search(r'ROW\(\)\s*\+\s*(\d+)', func_expr)
                    if match:
                        add_num = int(match.group(1))
                        result = row_num + add_num
                        self.add_result(f"        ROW()+{add_num} = {result}")
                        return result
                
                return row_num
            
            elif target_cell and func_type.upper() == 'COLUMN':
                col_letters = re.search(r'[A-Z]+', target_cell).group()
                col_num = 0
                for char in col_letters:
                    col_num = col_num * 26 + (ord(char) - ord('A') + 1)
                self.add_result(f"        COLUMN() context: {target_cell}, column: {col_num}")
                return col_num
            
            return f"{func_type}()"
        except:
            return f"{func_type}()"
    
    
    def resolve_external_reference_full(self, external_ref):
        """完整的外部引用解析"""
        try:
            # 簡化版本，返回工作表名稱
            if 'GDP' in external_ref:
                return "工作表2"  # 根據實際情況調整
            return "Unknown"
        except:
            return None
    
    def validate_inputs(self):
        """驗證輸入"""
        if not self.file_path_var.get():
            messagebox.showwarning("Warning", "Please select an Excel file")
            return False
        
        if not self.sheet_var.get():
            messagebox.showwarning("Warning", "Please select a worksheet")
            return False
        
        if not self.cell_var.get():
            messagebox.showwarning("Warning", "Please enter a cell address")
            return False
        
        return True
    
    def clear_results(self):
        """清除結果"""
        self.result_text.delete(1.0, tk.END)
    
    def add_result(self, message):
        """添加結果"""
        timestamp = time.strftime("%H:%M:%S")
        self.result_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.result_text.see(tk.END)
        self.root.update()
    
    def run(self):
        """運行GUI"""
        self.root.mainloop()

def main():
    """主函數"""
    if not HAS_OPENPYXL:
        print("ERROR: Need openpyxl")
        return
    
    app = UnifiedIndirectResolver()
    app.run()

if __name__ == "__main__":
    main()